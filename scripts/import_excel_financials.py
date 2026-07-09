#!/usr/bin/env python3
"""Import a simple Excel financial table into rows JSON for normalization."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required for Excel import") from exc


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--sheet")
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, data_only=True, read_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = ["" if value is None else value for value in row]
        if any(str(value).strip() for value in values):
            rows.append(values)
    args.out.write_text(json.dumps({"rows": rows}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"sheet": ws.title, "rows": len(rows), "out": str(args.out)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
